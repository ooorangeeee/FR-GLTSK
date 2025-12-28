# -*- coding: UTF-8 -*-
"""
@Author ：L.Gao
@Date ：2024/12/1 10:40 
"""
from FR_GLTSK.myfuzzytool import *
import openpyxl
import torch
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import KFold

# 【excel】载入电子表格，记录实验结果
xlsx_file = rf'./result/{"FR_reg_results"}.xlsx'
wb = openpyxl.load_workbook(xlsx_file)

# 实验重复次数
num_repeats = 5
num_folds = 10

# 数据集；数据集所在文件夹+数据集文件的后缀名
datasets_file = r'../data_pt/regression/{}.pt'
dataset_name_list = [
    "SINC",
    "Auto MPG",
    "RealEstate",
    "WSNs",
    # "Servo",
    # "QSAR",
    "NN5-110",
    "NNGC1-008",
]

# 超参数：[lr, num_fuzzy_set, epoch, coef_ant_lambda, coef_con_lambda, tau_fea, tau_rule, learning_rate_retrain, epoch_retrain]
para_dict = {
    13: [0.05, 30, 50, 0.0001, 0.0001, 0.05, 0.01, 0.2, 50],
    14: [0.05, 30, 50, 0.0002, 0.0002, 0.04, -0.01, 0.24, 50],
    15: [0.05, 30, 50, 0.0003, 0.0003, 0.03, -0.05, 0.28, 50],
    16: [0.05, 30, 50, 0.0004, 0.0004, 0.02, -0.1, 0.32, 50],
    17: [0.05, 30, 50, 0.0005, 0.0005, 0.01, -0.3, 0.36, 50],
}

for dataset_name in dataset_name_list:  # 对于数据集data_name
    # 【excel】调取表单，准备写入实验结果
    sh = wb[dataset_name]
    # 载入数据集
    dataset = torch.load(datasets_file.format(dataset_name))
    sample, target = dataset.sample, dataset.target

    for row_th, para_list in para_dict.items():  # 对于参数组合para_list
        # 获取单个参数
        learning_rate, num_fuzzy_set, epoch, coef_ant_lambda, coef_con_lambda, tau_fea, tau_rule, \
            learning_rate_retrain, epoch_retrain = para_list
        # 【excel】写入使用的参数
        sh['T{}'.format(row_th)] = fr'{para_list}'

        for t in range(num_repeats):  # 对于数据集data_name，对于参数组合para_str，第t次重复实验
            # 十折交叉验证
            test_rmse_k_fold = []
            num_selected_fea_k_fold = []
            num_extracted_rule_k_fold = []
            kf = KFold(n_splits=num_folds, shuffle=False)
            for k, (tra_ind, test_ind) in enumerate(kf.split(sample, target)):  # 十折交叉实验，当前一折
                # 获取训练集与测试集(验证集)
                tra_sam, tra_tar = sample[tra_ind], target[tra_ind]
                test_sam, test_tar = sample[test_ind], target[test_ind]

                # 数据归一化
                tra_sam_scaler = MinMaxScaler()
                tra_tar_scaler = MinMaxScaler()
                tra_sam = tra_sam_scaler.fit_transform(tra_sam)
                test_sam = tra_sam_scaler.transform(test_sam)
                tra_tar = tra_tar_scaler.fit_transform(tra_tar.reshape(-1, 1))
                test_tar = test_tar.unsqueeze(1)

                # ndarray转换为tensor
                tra_sam = torch.tensor(tra_sam)
                test_sam = torch.tensor(test_sam)
                tra_tar = torch.tensor(tra_tar)

                # 获取输入、输出维度
                input_dim = tra_sam.shape[1]
                output_dim = 1

                # 实例化模型
                model = IT2TSK(input_layer_size=input_dim, num_fuzzy=num_fuzzy_set, output_layer_size=output_dim,
                               tnorm="adasoftmin1", typereduced="KM", order="simpl_first", firing_normalized=True,
                               gl_fea_sel=True, gl_rule_extr=True, tau_fea=tau_fea)
                model.input_to_memf_layer.reinit_spread_zero()
                model.rule_to_typereduced_layer.reinit_con_zero()

                # 训练模型
                print(f"---------第{k+1}折实验开始---------")
                model.trained_param(("spread", "THEN"))
                train_full_batch(model, tra_sam, tra_tar, learning_rate, epoch, optim_type="Adam", gpu=True,
                                 task="regression", coef_ant_lambda=coef_ant_lambda, coef_con_lambda=coef_con_lambda)

                # 特征选择 * 规则提取
                selected_fea_ind = model.gl_fea_select(tau_fea)
                extracted_rule_ind = model.gl_rule_extract(tau_rule)
                model.prune_structure(selected_fea_ind, extracted_rule_ind)
                model.rule_to_typereduced_layer.gl_fea_sel = False
                model.gl_fea_sel, model.gl_rule_extr = False, False

                # 重训练
                model.reinit_ant_IT2FCM(tra_sam[:, selected_fea_ind])
                model.trained_param(('all',))
                train_full_batch(model, tra_sam[:, selected_fea_ind], tra_tar, learning_rate_retrain, epoch_retrain,
                                 optim_type="Adam", lr_sched=False, gpu=True, task="regression")

                # 测试模型
                _, _, test_rmse = test(model, test_sam[:, selected_fea_ind], test_tar, task="regression", pre_sca=tra_tar_scaler)

                # 记录单次十折交叉实验中每一折的结果
                test_rmse_k_fold.append(test_rmse)
                num_selected_fea_k_fold.append(len(selected_fea_ind))
                num_extracted_rule_k_fold.append(len(extracted_rule_ind))

                print(
                    '{}数据集，第{}组参数，第{}折数值实验，测试集精度：{:.4f}，被选特征数{}，被提取规则数{}'.format(
                        dataset_name, row_th, k + 1, test_rmse, len(selected_fea_ind), len(extracted_rule_ind)))
                print(f"---------第{k+1}折实验结束---------")

            # 计算单次重复实验中十折交叉的平均测试精度
            test_rmse_k_fold_mean = torch.Tensor(test_rmse_k_fold).mean()
            num_selected_fea_k_fold_mean = torch.Tensor(num_selected_fea_k_fold).mean()
            num_extracted_rule_k_fold_mean = torch.Tensor(num_extracted_rule_k_fold).mean()

            # 【excel】写入十折交叉验证平均精度、选择特征数、提取规则数
            sh.cell(row_th, t + 2).value = test_rmse_k_fold_mean.tolist()
            sh.cell(row_th, t + 2).number_format = '0.0000'
            sh.cell(row_th, t + num_repeats + 3).value = num_selected_fea_k_fold_mean.tolist()
            sh.cell(row_th, t + num_repeats + 3).number_format = '0.00'
            sh.cell(row_th, t + num_repeats * 2 + 4).value = num_extracted_rule_k_fold_mean.tolist()
            sh.cell(row_th, t + num_repeats * 2 + 4).number_format = '0.00'


        # 【excel】表格中计算num_repeats次实验的平均测试精度、选择特征数、提取规则数
        sh[f'G{row_th}'] = f'=AVERAGE(B{row_th}:F{row_th})'
        sh[f'G{row_th}'].number_format = '0.0000'
        sh[f'M{row_th}'] = f'=AVERAGE(H{row_th}:L{row_th})'
        sh[f'M{row_th}'].number_format = '0.00'
        sh[f'S{row_th}'] = f'=AVERAGE(N{row_th}:R{row_th})'
        sh[f'S{row_th}'].number_format = '0.00'
        # 【excel】保存
        wb.save(xlsx_file)

# 【excel】关闭
wb.close()



